import os
import json

import openpyxl
import pandas as pd
from bson import ObjectId
from typing import List

from app.enums.parsers import ParserStatus, ParserType
from app.repositories.user_parsers import UserParsersRepository
from app.schemas.parser import BaseParsersSchema
from app.services import parser
from app.database.mysql import MysqlManager


async def get_headers(keys, header_row: int, sheet):
    for col, key in enumerate(keys, start=1):
        sheet.cell(row=header_row, column=col, value=key)


async def get_keys(data):
    keys = set()
    for d in data:
        keys = keys.union(d)
    return keys


async def create_excel(data: list, name_of_excel: str) -> str:
    workbook = openpyxl.Workbook()
    data = json.loads(data)
    sheet = workbook.active
    keys = set()
    for json_data in data:
        local_keys = json_data.keys()
        for key in local_keys:
            keys.add(key)

    print(keys)

    for col, key in enumerate(keys, start=1):
        print(col)
        print(key)
        sheet.cell(row=1, column=col, value=key)  # type: ignore

    for row, json_data in enumerate(data, start=2):
        for col, key in enumerate(keys, start=1):
            for item in json_data.items():
                if key == item[0]:
                    sheet.cell(row=row, column=col, value=str(item[1]))  # type: ignore

    #await get_headers(keys, 1, sheet)
    #for row, obj in enumerate(data.values(), start=1 + 1):
        #for col, key in enumerate(keys, start=1):
            #print(obj)
            #value = str(obj)
            
    excel_file = f"{name_of_excel}.xlsx"
    workbook.save(excel_file)
    return excel_file


async def read_headers(excel_file):
    data_frame = pd.read_excel(excel_file)
    headers = data_frame.columns.tolist()
    return headers, data_frame


async def read_excel(excel_headers, df):
    row = 0
    first_row_values = {}
    result = []
    while row is not None:
        try:
            for header in excel_headers:
                first_row_values[header] = df[header].iloc[row]
            row += 1
            result.append(first_row_values)
        except IndexError:
            break
    return result


async def create_excel_file(base_id: str):
    query = await MysqlManager.cursor.execute(f"SELECT * FROM `parsers` WHERE `_id` = '{base_id}'")
    base = await MysqlManager.cursor.fetchall()
    base = base[0]
    data = base['parser_data']
    if base["parser_type"] == "Avito":
        file_name = "Avito"
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base["parser_type"] == "Yandex":
        file_name = "Yandex"
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base["parser_type"] == "Vk Groups":
        file_name = "Vk Groups"
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base["parser_type"] == "Vk Posts":
        file_name = "Vk Posts"
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name


async def delete_file(file_path: str):
    os.remove(file_path)


async def import_excel(owner_id: str, excel_file, parser_type: str, filters: List[str]):
    headers, data_frame = await read_headers(excel_file)
    excel_data = await read_excel(headers, data_frame)
    parser_data = BaseParsersSchema(
        parser_type=parser_type,
        owner_id=owner_id,
        status=ParserStatus.parsed,
        parser_data=excel_data,
        filters=filters
    )
    user_parser = await parser.create_base(parser_data)
    return user_parser
